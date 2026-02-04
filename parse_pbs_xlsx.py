#!/usr/bin/env python3
"""
Parse PBS XLSX file to extract state-level prescribing data
Uses the main "Date of Supply" report (98MB XLSX)
"""
import sys
from openpyxl import load_workbook
from collections import defaultdict
import csv

def analyze_xlsx_structure(xlsx_path):
    """Analyze XLSX file structure to understand the data"""
    print("="*80)
    print("PBS XLSX File Structure Analysis")
    print("="*80)
    
    print(f"\nLoading workbook: {xlsx_path}")
    print("(This may take 1-2 minutes for large files...)")
    
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    
    print(f"\n✓ Workbook loaded")
    print(f"Sheet names: {wb.sheetnames}")
    
    # Analyze first sheet
    first_sheet = wb[wb.sheetnames[0]]
    
    print(f"\nAnalyzing sheet: {first_sheet.title}")
    
    # Get headers (first row)
    headers = []
    for cell in first_sheet[1]:
        headers.append(cell.value)
    
    print(f"\nColumns ({len(headers)}):")
    for i, header in enumerate(headers, 1):
        print(f"  {i}. {header}")
    
    # Sample first few data rows
    print(f"\nSample data (first 3 rows):")
    for row_idx, row in enumerate(first_sheet.iter_rows(min_row=2, max_row=4), 1):
        print(f"\nRow {row_idx}:")
        for i, cell in enumerate(row):
            if i < len(headers):
                print(f"  {headers[i]}: {cell.value}")
    
    return wb, headers

def extract_metformin_state_data(xlsx_path, drug_codes):
    """
    Extract state-level metformin data from PBS XLSX
    
    Args:
        xlsx_path: Path to PBS XLSX file
        drug_codes: Set of PBS item codes for metformin
    """
    print("\n" + "="*80)
    print("Extracting Metformin State Data")
    print("="*80)
    
    print(f"\nLoading workbook...")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = wb.active
    
    # Get headers
    headers = [cell.value for cell in sheet[1]]
    
    # Find column indices
    try:
        month_col = headers.index('MONTH_OF_SUPPLY') if 'MONTH_OF_SUPPLY' in headers else headers.index('Supply Month')
        item_col = headers.index('ITEM_CODE') if 'ITEM_CODE' in headers else headers.index('PBS Item Code')
        state_col = None
        
        # Look for state column (various possible names)
        for possible_state in ['STATE', 'State', 'STATE_CODE', 'State Code', 'JURISDICTION']:
            if possible_state in headers:
                state_col = headers.index(possible_state)
                break
        
        if state_col is None:
            print(f"\n⚠️  WARNING: No state column found!")
            print(f"Available columns: {headers}")
            return None
        
        rx_col = headers.index('PRESCRIPTIONS') if 'PRESCRIPTIONS' in headers else headers.index('Scripts')
        cost_col = headers.index('TOTAL_COST') if 'TOTAL_COST' in headers else headers.index('Total Cost')
        
        print(f"\n✓ Found columns:")
        print(f"  Month: {headers[month_col]} (col {month_col})")
        print(f"  Item: {headers[item_col]} (col {item_col})")
        print(f"  State: {headers[state_col]} (col {state_col})")
        print(f"  Prescriptions: {headers[rx_col]} (col {rx_col})")
        print(f"  Cost: {headers[cost_col]} (col {cost_col})")
        
    except ValueError as e:
        print(f"\n❌ Error finding columns: {e}")
        print(f"Available columns: {headers}")
        return None
    
    # Aggregate data by state and month
    print(f"\nProcessing rows (this may take a few minutes)...")
    state_month_data = defaultdict(lambda: defaultdict(lambda: {'rx': 0, 'cost': 0.0}))
    
    row_count = 0
    metformin_rows = 0
    
    for row in sheet.iter_rows(min_row=2):
        row_count += 1
        
        if row_count % 100000 == 0:
            print(f"  Processed {row_count:,} rows, found {metformin_rows:,} metformin records...")
        
        try:
            item_code = str(row[item_col].value).strip() if row[item_col].value else ""
            
            if item_code in drug_codes:
                month = str(row[month_col].value)
                state = str(row[state_col].value).strip() if row[state_col].value else "UNK"
                rx = int(row[rx_col].value) if row[rx_col].value else 0
                cost = float(row[cost_col].value) if row[cost_col].value else 0.0
                
                state_month_data[state][month]['rx'] += rx
                state_month_data[state][month]['cost'] += cost
                metformin_rows += 1
                
        except Exception as e:
            # Skip problematic rows
            continue
    
    print(f"\n✓ Processing complete")
    print(f"  Total rows processed: {row_count:,}")
    print(f"  Metformin records found: {metformin_rows:,}")
    print(f"  States found: {len(state_month_data)}")
    
    return state_month_data

def export_results(state_month_data, output_file='pbs_metformin_by_state.csv'):
    """Export aggregated data to CSV"""
    print(f"\n\nExporting results to {output_file}...")
    
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['State', 'Month', 'Prescriptions', 'Cost_AUD'])
        
        for state in sorted(state_month_data.keys()):
            for month in sorted(state_month_data[state].keys()):
                data = state_month_data[state][month]
                writer.writerow([state, month, data['rx'], f"{data['cost']:.2f}"])
    
    print(f"✓ Exported {sum(len(months) for months in state_month_data.values())} rows")
    
    # Print summary
    print(f"\n\nState Summary:")
    for state in sorted(state_month_data.keys()):
        total_rx = sum(d['rx'] for d in state_month_data[state].values())
        total_cost = sum(d['cost'] for d in state_month_data[state].values())
        print(f"  {state}: {total_rx:,} prescriptions, ${total_cost:,.2f} AUD")

def main():
    xlsx_file = 'pbs_data/pbs_jul2021_oct2025.xlsx'
    
    # Load metformin codes from drug map
    print("Loading metformin PBS item codes...")
    metformin_codes = set()
    with open('pbs_data/pbs_item_drug_map.csv', 'r', encoding='latin-1') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if 'METFORMIN' in row['DRUG_NAME'].upper():
                metformin_codes.add(row['ITEM_CODE'])
    
    print(f"✓ Found {len(metformin_codes)} metformin item codes")
    
    # First, analyze structure
    print(f"\nStep 1: Analyzing XLSX structure...")
    try:
        wb, headers = analyze_xlsx_structure(xlsx_file)
        wb.close()
    except FileNotFoundError:
        print(f"\n❌ File not found: {xlsx_file}")
        print(f"Waiting for download to complete...")
        return
    except Exception as e:
        print(f"\n❌ Error analyzing file: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Extract data
    print(f"\nStep 2: Extracting metformin state data...")
    state_data = extract_metformin_state_data(xlsx_file, metformin_codes)
    
    if state_data:
        # Export results
        export_results(state_data, 'pbs_data/metformin_by_state.csv')
        
        print("\n" + "="*80)
        print("✅ SUCCESS")
        print("="*80)
        print(f"Real PBS state-level data extracted and saved!")
    else:
        print("\n⚠️  Could not extract state data from this file")

if __name__ == "__main__":
    main()
